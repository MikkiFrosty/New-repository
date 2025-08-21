from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from io import TextIOWrapper
from pathlib import Path

BASE = Path(__file__).resolve().parent

def test_zip():
    with ZipFile('zip_file.zip', 'w') as zipf:
        zipf.write(BASE / 'Test.csv', arcname='Test.csv')
        zipf.write(BASE / 'Test.xlsx', arcname='Test.xlsx')
        zipf.write(BASE / 'Test.pdf', arcname='Test.pdf')

    with ZipFile('zip_file.zip', 'r') as zipf:
        with zipf.open('Test.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            print(reader.pages[0].extract_text())
            assert "Test" in reader.pages[0].extract_text()

    with ZipFile('zip_file.zip', 'r') as zipf:
        with zipf.open('Test.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            print(sheet.cell(row=2, column=1).value)
            assert 'StartDate' in sheet.cell(row=1, column=2).value

    with ZipFile('zip_file.zip', 'r') as zipf:
        with zipf.open('Test.csv') as csv_file:
            text_file = TextIOWrapper(csv_file, encoding='utf-8-sig')
            content = text_file.read()
            print(content)
            assert ".com" in content